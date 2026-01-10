import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from scipy.optimize import curve_fit

# Corrected function for RLC resonance
def frequence_var(L, C):
    return 1/np.sqrt(L*C)

# Load data from Excel
wb = openpyxl.load_workbook('1_data.xlsx')
wb = wb.active

data_max = 12
row_1 = 3
col_1 = 2

# Extracting data from Excel
L = np.array([wb.cell(row=row_1, column=i).value for i in range(col_1, data_max)])
T = np.array([wb.cell(row=row_1+4, column=i).value for i in range(col_1, data_max)])
dT = np.array([wb.cell(row=row_1+6, column=i).value for i in range(col_1, data_max)])

# Convert period to angular frequency
w = [2 * np.pi / T[i] for i in range(len(T))]
dw = [2 * np.pi*dT[i] / T[i]/T[i] for i in range(len(dT))]

# Initial guesses for fitting parameters
C_guess = 1e-7

# Fit the data
parameters, covariance = curve_fit(frequence_var, L, w, p0=[C_guess])

# Extract fitted parameters
C_fit = parameters

# Print fitted parameters
print(f"Fitted Parameters:\nC = {C_fit} F")

# Plot experimental data with error bars
plt.errorbar(L, w, xerr=0.01, fmt='k.', label="Datos")

# Plot fitted curve
L_fit = np.linspace(min(L), max(L), 500)
plt.plot(L_fit, frequence_var(L_fit,C_fit), 'r-', label="Ajuste")
plt.title("Frecuencia de resonancia para variación de L")

# Labels and legend
plt.xlabel(r"$L$ (H)")
plt.ylabel(r"$\omega_0$ (rad/s)")
plt.legend()
plt.grid()
plt.show()