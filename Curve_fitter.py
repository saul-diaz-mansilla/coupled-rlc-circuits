#

import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from scipy.optimize import curve_fit

def Lorentzian(x,x0,gamma):
    return 1/(np.pi*gamma*(1+((x-x0)/gamma)**2))

wb = openpyxl.load_workbook('Datos_semana_1.xlsx')
wb = wb.active
data_max = 19
row_1 = 17
col_1 = 2

f = [wb.cell(row=row_1,column=i).value for i in range(col_1,data_max)]
df = [wb.cell(row=row_1+1,column=i).value for i in range(col_1,data_max)]
Vpp = [wb.cell(row=row_1+2,column=i).value for i in range(col_1,data_max)]
dV = [wb.cell(row=row_1+3,column=i).value for i in range(col_1,data_max)]
t = [wb.cell(row=row_1+4,column=i).value for i in range(col_1,data_max)]
dt = [wb.cell(row=row_1+5,column=i).value for i in range(col_1,data_max)]
delta = [2*np.pi*f[i]*t[i]*1e-3 for i in range(1,len(f))]
ddelta = [2*np.pi*(df[i]*t[i]+f[i]*dt[i]) for i in range(1,len(f))]

parameters,err = curve_fit(Lorentzian,2*np.pi*np.array(f),np.array(Vpp)/2)
print(parameters[0])

plt.errorbar(2*np.pi*np.array(f),np.array(Vpp)/2,np.array(dV)/2,2*np.pi*np.array(df),'k.')
plt.plot(2*np.pi*np.array(f),Lorentzian(2*np.pi*np.array(f),parameters[0],parameters[1]),'r-')
plt.xlabel(r"$\omega$ (rad/s)")
plt.ylabel(r"$V_{máx}$ (V)")
plt.show()