# Bode Capacitivo

import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from scipy.optimize import curve_fit

def G(w,wn,beta):
    return 1/np.sqrt((w**2-wn**2)**2+(2*beta*w)**2)
def phi(w,wn,beta):
    val = np.arctan2(-2*beta*w,w**2-wn**2)
    return val + np.pi*(val<0)

def A1(w,beta,V0):
    G1 = G(w,w1,beta)
    G2 = G(w,w2,beta)
    phi1 = phi(w,w1,beta)
    phi2 = phi(w,w2,beta)
    return V0/2/L/C*np.sqrt(G1**2+G2**2+2*G1*G2*np.cos(phi1-phi2))

def A2(w,beta,V0):
    G1 = G(w,w1,beta)
    G2 = G(w,w2,beta)
    phi1 = phi(w,w1,beta)
    phi2 = phi(w,w2,beta)
    return V0/2/L/C*np.sqrt(G1**2+G2**2-2*G1*G2*np.cos(phi1-phi2))

def theta1(w,R,L,C,C3,V0):
    G1 = G(w,np.sqrt(1/L/C),R/2/L)
    G2 = G(w,np.sqrt(1/L/C+2/L/C3),R/2/L)
    phi1 = phi(w,np.sqrt(1/L/C),R/2/L)
    phi2 = phi(w,np.sqrt(1/L/C+2/L/C3),R/2/L)
    val = np.arctan2(G1*np.sin(phi1)+G2*np.sin(phi2),G1*np.cos(phi1)+G2*np.cos(phi2))
    return val

def theta2(w,R,L,C,C3,V0):
    G1 = G(w,np.sqrt(1/L/C),R/2/L)
    G2 = G(w,np.sqrt(1/L/C+2/L/C3),R/2/L)
    phi1 = phi(w,np.sqrt(1/L/C),R/2/L)
    phi2 = phi(w,np.sqrt(1/L/C+2/L/C3),R/2/L)
    val = np.arctan2(-G1*np.sin(phi1)+G2*np.sin(phi2),-G1*np.cos(phi1)+G2*np.cos(phi2))
    return val

wb = openpyxl.load_workbook(r"C:\Users\Usuario\Documents\Documentos\Universidad\Tecnicas experimentales II\Proyecto\Datos_semana_2.xlsx")
wb = wb.active

data_max = 23
col_1 = 2
row_1 = 19

# Extracting data from Excel
f = np.array([wb.cell(row=row_1, column=i).value for i in range(col_1, data_max)])
df = np.array([wb.cell(row=row_1+1, column=i).value for i in range(col_1, data_max)])
Vpp = np.array([wb.cell(row=row_1+2, column=i).value for i in range(col_1, data_max)])
dV = np.array([wb.cell(row=row_1+3, column=i).value for i in range(col_1, data_max)])
Vmax = np.array([wb.cell(row=row_1+4, column=i).value for i in range(col_1, data_max)])
Vmin = np.array([wb.cell(row=row_1+5, column=i).value for i in range(col_1, data_max)])
dVm = np.array([wb.cell(row=row_1+6, column=i).value for i in range(col_1, data_max)])
T = np.array([wb.cell(row=row_1+7, column=i).value for i in range(col_1, data_max)])
dT = np.array([wb.cell(row=row_1+8, column=i).value for i in range(col_1, data_max)])
T = T*1e-6
dT = dT*1e-6


VDC = Vmax+Vmin
w = 2*np.pi*f
dw = 2*np.pi*df
delta = 2*np.pi*f*T
ddelta = 2*np.pi*(df*T+f*dT)

# Guesses
L = wb.cell(row=14,column=2).value
C = wb.cell(row=15,column=2).value
R = wb.cell(row=16,column=2).value
R = 1700
C3 = wb.cell(row=17,column=2).value
C3 = 1e-8
V0 = wb.cell(row=18,column=2).value/2
V0 = 20.6/2

w1 = 1/np.sqrt(L*C)
w2 = np.sqrt(1/L/C+2/L/C3)

print([R, L, C, C3, V0])

figure1 = plt.figure()
fig1 = figure1.add_subplot(111)
w_fit = np.linspace(14, 5000, 500)

fig1.plot(w_fit, A1(w_fit*2*np.pi, R/2/L, V0), label="Teórico 1")
fig1.plot(w_fit, A2(w_fit*2*np.pi, R/2/L, V0), label="Teórico 2")

# Labels and legend
fig1.set_title("Resupuesta del sistema a ondas sinusolidales")
fig1.set_xlabel(r"$f$ (Hz)")
# fig1.set_xlabel(r"$\omega$ (rad/s)")
fig1.set_ylabel(r"$V_{máx}$ (V)")
fig1.legend()
fig1.grid()

plt.show()

# plt.plot(w_fit, theta1(w_fit,R,L,C,C3,V0), label="Teórico 1")
# plt.plot(w_fit, theta2(w_fit,R,L,C,C3,V0), label="Teórico 2")
# plt.legend()
# plt.grid()
# plt.show()
