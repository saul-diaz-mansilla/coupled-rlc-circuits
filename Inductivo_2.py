# Bode Capacitivo

import numpy as np
from numpy import pi as π
import matplotlib.pyplot as plt
import openpyxl
from scipy.optimize import curve_fit
import scipy.integrate as spi
import scipy.special as sp

μ_0 = 4*π*1e-7
N = 500
R1 = 10.06e-2
R2 = 11.37e-2

def M(d):
    return μ_0*π*N**2*R/2/(1+(R/d)**2)**1.5

# Function for mutual inductance
# def M(d):
#     """
#     Compute the mutual inductance M between two coaxial circular coils using elliptic integrals.
    
#     Parameters:
#     R1 : float  -> Radius of first coil (m)
#     R2 : float  -> Radius of second coil (m)
#     d  : float  -> Distance between the two coils (m)
#     N : int    -> Number of turns in first coil
#     N : int    -> Number of turns in second coil
    
#     Returns:
#     M  : float  -> Mutual inductance (H)
#     """
    
#     def integrand(r):
#         k2 = 4 * R1 * r / ((R1 + r) ** 2 + d ** 2)  # Argument for elliptic integral
#         k = np.sqrt(k2)
        
#         if k2 >= 1:  # Prevent numerical issues
#             return 0
        
#         Kk = sp.ellipk(k2)  # Complete elliptic integral of the first kind
#         Ek = sp.ellipe(k2)  # Complete elliptic integral of the second kind
        
#         term1 = Kk
#         term2 = ((R1**2 - r**2 - d**2) / ((R1 - r)**2 + d**2)) * Ek
        
#         return r * (term1 + term2) / np.sqrt((R1 + r)**2 + d**2)

#     # Perform numerical integration
#     integral_result, _ = spi.quad(integrand, 0, R2)

#     # Compute mutual inductance
#     M = (μ_0 * N * N * R1**2 / (2 * np.pi)) * integral_result
    
#     return M

def G(w,wn,beta):
    return 1/np.sqrt((w**2-wn**2)**2+(2*beta*w)**2)
def phi(w,wn,beta):
    val = np.arctan(-2*beta*w,w**2-wn**2)
    return val + π*(val<0)

def A1(w,beta1,beta2,V0):
    G1 = G(w,w1,beta1)
    G2 = G(w,w2,beta2)
    phi1 = phi(w,w1,beta1)
    phi2 = phi(w,w2,beta2)
    return V0/2/C*np.sqrt(G1**2/(L+M(d))**2+G2**2/(L-M(d))**2+2*G1*G2/(L**2-M(d)**2)*np.cos(phi1-phi2))

def A2(w,beta1,beta2,V0):
    G1 = G(w,w1,beta1)
    G2 = G(w,w2,beta2)
    phi1 = phi(w,w1,beta1)
    phi2 = phi(w,w2,beta2)
    return V0/2/C*np.sqrt(G1**2/(L+M(d))**2+G2**2/(L-M(d))**2-2*G1*G2/(L**2-M(d)**2)*np.cos(phi1-phi2))

def theta1(w,R,L,C,C3,V0):
    G1 = G(w,np.sqrt(1/L/C),R/2/L)
    G2 = G(w,np.sqrt(1/L/C+2/L/C3),R/2/L)
    phi1 = phi(w,np.sqrt(1/L/C),R/2/L)
    phi2 = phi(w,np.sqrt(1/L/C+2/L/C3),R/2/L)
    val = np.arctaN(G1*np.sin(phi1)+G2*np.sin(phi2),G1*np.cos(phi1)+G2*np.cos(phi2))
    return val + π*(val<0)

def theta2(w,R,L,C,C3,V0):
    G1 = G(w,np.sqrt(1/L/C),R/2/L)
    G2 = G(w,np.sqrt(1/L/C+2/L/C3),R/2/L)
    phi1 = phi(w,np.sqrt(1/L/C),R/2/L)
    phi2 = phi(w,np.sqrt(1/L/C+2/L/C3),R/2/L)
    val = np.arctaN(-G1*np.sin(phi1)+G2*np.sin(phi2),-G1*np.cos(phi1)+G2*np.cos(phi2))
    return val + π*(val<0)

wb = openpyxl.load_workbook(r"C:\Users\Usuario\Documents\Documentos\Universidad\Tecnicas experimentales II\Proyecto\Datos_semana_2.xlsx")
wb = wb.active

data_max = 32
col_1 = 2
row_1 = 51

# Extracting data from Excel
f = np.array([wb.cell(row=row_1, column=i).value for i in range(col_1, data_max)])
df = np.array([wb.cell(row=row_1+1, column=i).value for i in range(col_1, data_max)])
Vpp = np.array([wb.cell(row=row_1+2, column=i).value for i in range(col_1, data_max)])
dV = np.array([wb.cell(row=row_1+3, column=i).value for i in range(col_1, data_max)])
T = np.array([wb.cell(row=row_1+5, column=i).value for i in range(col_1, data_max)])
dT = np.array([wb.cell(row=row_1+6, column=i).value for i in range(col_1, data_max)])
T = T*1e-6
dT = dT*1e-6


w = 2*π*f
dw = 2*π*df
delta = 2*π*f*T
ddelta = 2*π*(df*T+f*dT)

# Guesses
L = 25e-3
C = wb.cell(row=48,column=2).value
R = wb.cell(row=49,column=2).value
d = wb.cell(row=50,column=2).value
V0 = 19.2

w1 = 1/np.sqrt(C*(L+M(d)))
w2 = 1/np.sqrt(C*(L-M(d)))
print(w1)
print(w2)
beta1 = R/2/(L+M(d))
beta2 = R/2/(L-M(d))

figure1 = plt.figure()
fig1 = figure1.add_subplot(111)
# figure2 = plt.figure()
# fig2 = figure2.add_subplot(111)
w_fit = np.linspace(min(w), max(w), 500)

parameters, covariance = curve_fit(A1, w, Vpp/2, p0=[beta1, beta2, V0])  

# Plot experimental data with error bars
fig1.errorbar(w, Vpp/2, yerr=dV, xerr=dw, fmt='k.', label=r"Datos $V_1$")
fig1.plot(w_fit, A1(w_fit, parameters[0], parameters[1], parameters[2]), 'r-', label=r"Ajuste $V_1$")

plt.show()