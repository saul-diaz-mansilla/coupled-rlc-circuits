import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from scipy.optimize import curve_fit

# Define the theoretical time shift function
def time_shift(w, R, L, C):
    return (1 / w) * np.arctan((w * L - 1 / (w * C)) / R)

# Load data from Excel
wb = openpyxl.load_workbook('Datos_semana_1.xlsx')
wb = wb.active

data_max = 19
row_1 = 17
col_1 = 2

# Extract frequency and time shift data from Excel
f = np.array([wb.cell(row=row_1, column=i).value for i in range(col_1, data_max)])
df = np.array([wb.cell(row=row_1+1, column=i).value for i in range(col_1, data_max)])
t = np.array([wb.cell(row=row_1+4, column=i).value for i in range(col_1, data_max)])  # Experimental time shift
dt = np.array([wb.cell(row=row_1+5, column=i).value for i in range(col_1, data_max)])  # Error in time shift

# Convert frequency to angular frequency
w = 2 * np.pi * f
dw = 2 * np.pi * df

# Initial guesses for R, L, C
R_guess = 100  # Ohms
L_guess = 0.001  # Henry
C_guess = 1e-6  # Farads

# Fit the experimental time shift to the theoretical function
params, covariance = curve_fit(time_shift, w, t, p0=[R_guess, L_guess, C_guess])

# Extract fitted parameters
R_fit, L_fit, C_fit = params

# Print fitted values
print(f"Fitted Parameters:\nR = {R_fit:.4f} Ω, L = {L_fit:.6f} H, C = {C_fit:.9f} F")

# Generate smooth fit curve
w_fit = np.linspace(min(w), max(w), 500)
t_fit = time_shift(w_fit, R_fit, L_fit, C_fit)

# Plot experimental data with error bars
plt.errorbar(w, t, yerr=dt, xerr=dw, fmt='k.', label="Experimental Data")

# Plot fitted curve
plt.plot(w_fit, t_fit, 'r-', label="Time Shift Fit")

# Labels and legend
plt.xlabel(r"$\omega$ (rad/s)")
plt.ylabel(r"Time Shift $t$ (s)")
plt.legend()
plt.grid()
plt.show()
