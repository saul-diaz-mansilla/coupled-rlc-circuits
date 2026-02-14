import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from scipy.optimize import curve_fit

# Corrected function for RLC resonance
def RLC_Resonance(w, A, w0, gamma):
    return A / np.sqrt((w0**2 - w**2)**2 + (gamma * w)**2)

# Load data from Excel
wb = openpyxl.load_workbook("2_data.xlsx")
wb = wb.active

data_max = 19
row_1 = 29
col_1 = 2

# Extracting data from Excel
f = np.array([wb.cell(row=row_1, column=i).value for i in range(col_1, data_max)])
df = np.array([wb.cell(row=row_1+1, column=i).value for i in range(col_1, data_max)])
Vpp = np.array([wb.cell(row=row_1+2, column=i).value for i in range(col_1, data_max)])
dV = np.array([wb.cell(row=row_1+3, column=i).value for i in range(col_1, data_max)])

# Convert frequency to angular frequency
w = 2 * np.pi * f
dw = 2 * np.pi * df

# Initial guesses for fitting parameters
A_guess = max(Vpp)   # Peak amplitude
w0_guess = w[np.argmax(Vpp)]  # Peak frequency
gamma_guess = (w[-1] - w[0]) / 10  # Rough estimate of width

# Fit the data
parameters, covariance = curve_fit(RLC_Resonance, w, Vpp, p0=[A_guess, w0_guess, gamma_guess])

# Extract fitted parameters
A_fit, w0_fit, gamma_fit = parameters

# Print fitted parameters
print(f"Fitted Parameters:\nA = {A_fit:.4f} V/s^2, ω₀ = {w0_fit:.4f} rad/s, γ = {gamma_fit:.4f} rad/s")

# Plot experimental data with error bars
plt.errorbar(w, Vpp, yerr=dV, xerr=dw, fmt='k.', label="Datos")

# Plot fitted curve
w_fit = np.linspace(min(w), max(w), 500)
plt.plot(w_fit, RLC_Resonance(w_fit, A_fit, w0_fit, gamma_fit), 'r-', label="Lorentziana")
plt.title("Amplitud de las oscilaciones para frecuencias de entrada")

# Labels and legend
plt.xlabel(r"$\omega$ (rad/s)")
plt.ylabel(r"$V_{máx}$ (V)")
plt.legend()
plt.grid()
plt.show()