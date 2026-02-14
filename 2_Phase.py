import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from scipy.optimize import curve_fit

# Define the phase shift function based on the RLC circuit
def phase_shift(w, A, w0, gamma):
    return np.arctan(-gamma*w0/(w**2-w0**2)) + np.pi*(w>w0)

# Load data from Excel
wb = openpyxl.load_workbook("2_data.xlsx")
wb = wb.active

data_max = 17
row_1 = 29
col_1 = 2

# Extract frequency and phase shift data from Excel
f = np.array([wb.cell(row=row_1, column=i).value for i in range(col_1, data_max)])
df = np.array([wb.cell(row=row_1+1, column=i).value for i in range(col_1, data_max)])
t = np.array([wb.cell(row=row_1+4, column=i).value for i in range(col_1, data_max)])
t = np.array(t)*1e-3
dt = np.array([wb.cell(row=row_1+5, column=i).value for i in range(col_1, data_max)])
dt = np.array(dt)*1e-3

# Compute the experimental phase shift using the incorrect method (for comparison)
delta_exp = 2 * np.pi * f * t
ddelta_exp = 2 * np.pi * (df * t + f * dt)

# Convert frequency to angular frequency
w = 2 * np.pi * f
dw = 2 * np.pi * df

# Initial guesses for R, L, C
A_guess = 199907945.6154  # Ohms
w0_guess = 3065.6491  # Henry
gamma_guess = 920.3916  # Farads

# Fit the phase shift data using curve_fit
params, covariance = curve_fit(phase_shift, w, delta_exp, p0=[A_guess, w0_guess, gamma_guess])

# Extract fitted parameters
A_fit, w0_fit, gamma_fit = params

# Print fitted values
print(f"Fitted Parameters:\nA = {A_fit:.4f} V/s^2, w0 = {w0_fit:.6f} rad/s, gamma = {gamma_fit:.9f} s^-1")

# Generate smooth fit curve
w_fit = np.linspace(min(w), max(w), 500)
delta_fit = phase_shift(w_fit, A_fit, w0_fit, gamma_fit)

# Plot experimental data with error bars
plt.errorbar(w, delta_exp, yerr=ddelta_exp, xerr=dw, fmt='k.', label="Datos experimentales")

# Plot fitted curve
plt.plot(w_fit, delta_fit, 'r-', label="Ajuste")
plt.title("Desfase entre la señal de entrada y en el condensador")

# Labels and legend
plt.xlabel(r"$\omega$ (rad/s)")
plt.ylabel(r"Desfase $\delta$ (rad)")
plt.legend()
plt.grid()
plt.show()
